#!/usr/bin/env python3

"""spreadsheet auto-fill

This script allows the user to automate the process to extract especific
information from a quotation to fill an .xlsx file.

This script is required to be located in the same directory as the .pdf
files as well as the main .xslx

This file contains the main function.

"""

import PyPDF2
import openpyxl
from pdf_regex_fun import pdf_regex
from recent_pdf import recent_pdf
from fill_excel import fill_excel
import sys

# Getting the most recent PDF file from the current directory
pdf_list = recent_pdf()
if len(pdf_list) == 0:
    print("Sorry, inside this directory there are no PDF files")
    sys.exit(1)

# opening excel file and setting the worksheet
try:
    pipeline_file = openpyxl.load_workbook('Pipeline.xlsx')
    work_sheet = pipeline_file['Hoja1']
except FileNotFoundError:
    print("Sorry the file Pipeline.xlsx doesn't exist inside this folder")
    sys.exit(1)
except:
    print("Unexpected error:", sys.exc_info()[0])
    sys.exit(1)

# Getting the first empty row
empty_cell = work_sheet.max_row + 1

# Opening the PDF file and saving the text content in pdf_text variable
for recent_pdf_file in pdf_list:
    empty_cell = work_sheet.max_row + 1
    with open(recent_pdf_file, 'rb') as pdf_file:
        file_reader = PyPDF2.PdfFileReader(pdf_file)
        file_page = file_reader.getPage(0)
        pdf_text = file_page.extractText()
    if not pdf_text:
        print("no data in file " + recent_pdf_file)
        continue

# Obtaining the dictionary with the information extracted from pdf
    dict_values = pdf_regex(pdf_text)
    if not dict_values:
        print("Couldn't find the required content")
        sys.exit(1)

    # Ideinfying if the quotation was already inserted in the .xlsx file
    i = 0
    for row in work_sheet.values:
        if (row[0] == dict_values['quot_number']):
            print("Quote {} already inserted".format(
                dict_values['quot_number']))
            i += 1
    # Applying style and value to each cell from A# to Q#
    if (i == 0):
        fill_excel(dict_values, work_sheet, empty_cell)

pipeline_file.save('Pipeline.xlsx')
pipeline_file.close()


# Make de proggram portable = "pip install pyinstaller",
# "got to direcotry -> app", "use pyinstaller --windowed
# --onefile --icon=./<icon.ico> <app_name>"
# Generate windows alert for every error
